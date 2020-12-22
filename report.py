from openpyxl import Workbook, load_workbook

class ExelPlain:

    @staticmethod
    def load_xlsx(path):
        return load_workbook(path)

    def __init__(self):
        self.document = self.create_document()
        self.index = 0

    def create_document(self):
        wb = Workbook()
        # ws = wb.active
        return wb

    def create_sheet(self, title):
        ws = self.document.create_sheet(title, self.index)
        ws['A1'] = "Заголовок"
        ws['B1'] = "Краткая новость"
        ws['C1'] = "Полная новость"
        ws['D1'] = "author"
        ws['E1'] = "released"
        ws['F1'] = "released-data"
        ws['G1'] = "label"
        ws['H1'] = "tracks"
        ws['I1'] = "duration"
        ws['J1'] = "album"
        ws['K1'] = "genre"
        return self.document


if __name__ == "__main__":
    PATH_FILE = "music.xlsx"
    try:
        wb = ExelPlain.load_xlsx(PATH_FILE)
    except FileNotFoundError:
        wb = ExelPlain()
        wb = wb.create_sheet("Music")

    sheet_ranges = wb["Music"]
    sheet_ranges.append(['1', '2222', '3', '4', '21', '12', '123', 'asd', 'b'])
    wb.save(PATH_FILE)


    # title = "Test title"
    # attrs = [("a1","b2","c3","d4")]
    # one = ExelPlain()
    # one.create_sheet(title, attrs)
    # # one.create_sheet(cor1)
    # one.save("temp.xlsx")
